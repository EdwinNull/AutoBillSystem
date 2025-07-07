#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出工具模块
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from config.settings import REPORT_DIR
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExportUtils:
    """导出工具类"""
    
    @staticmethod
    def export_to_csv(data, filename, headers=None):
        """导出数据到CSV文件"""
        try:
            # 确保报表目录存在
            REPORT_DIR.mkdir(exist_ok=True)
            
            filepath = REPORT_DIR / filename
            
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if not data:
                    return str(filepath)
                
                # 如果没有提供headers，从第一行数据获取
                if not headers:
                    if isinstance(data[0], dict):
                        headers = list(data[0].keys())
                    else:
                        headers = [f'列{i+1}' for i in range(len(data[0]))]
                
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                
                for row in data:
                    if isinstance(row, dict):
                        writer.writerow(row)
                    else:
                        # 如果是列表或元组，转换为字典
                        row_dict = {headers[i]: row[i] for i in range(len(row))}
                        writer.writerow(row_dict)
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"CSV导出失败: {e}")
    
    @staticmethod
    def export_to_json(data, filename, indent=2):
        """导出数据到JSON文件"""
        try:
            # 确保报表目录存在
            REPORT_DIR.mkdir(exist_ok=True)
            
            filepath = REPORT_DIR / filename
            
            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, ensure_ascii=False, indent=indent, default=str)
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"JSON导出失败: {e}")
    
    @staticmethod
    def export_to_txt(data, filename, separator='\t'):
        """导出数据到文本文件"""
        try:
            # 确保报表目录存在
            REPORT_DIR.mkdir(exist_ok=True)
            
            filepath = REPORT_DIR / filename
            
            with open(filepath, 'w', encoding='utf-8') as txtfile:
                if isinstance(data, list):
                    for row in data:
                        if isinstance(row, dict):
                            line = separator.join(str(v) for v in row.values())
                        elif isinstance(row, (list, tuple)):
                            line = separator.join(str(v) for v in row)
                        else:
                            line = str(row)
                        txtfile.write(line + '\n')
                else:
                    txtfile.write(str(data))
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"文本导出失败: {e}")
    
    @staticmethod
    def export_to_excel(data, filename, headers=None, sheet_name='Sheet1'):
        """导出数据到Excel文件"""
        if not EXCEL_AVAILABLE:
            raise Exception("Excel导出功能需要安装openpyxl库: pip install openpyxl")
        
        try:
            # 确保报表目录存在
            REPORT_DIR.mkdir(exist_ok=True)
            
            filepath = REPORT_DIR / filename
            
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 设置标题样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入表头
            if headers:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                start_row = 2
            else:
                start_row = 1
            
            # 写入数据
            if isinstance(data, list):
                for row_idx, row_data in enumerate(data, start_row):
                    if isinstance(row_data, dict):
                        if headers:
                            for col_idx, header in enumerate(headers, 1):
                                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                        else:
                            for col_idx, value in enumerate(row_data.values(), 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                    elif isinstance(row_data, (list, tuple)):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        ws.cell(row=row_idx, column=1, value=str(row_data))
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(filepath)
            return str(filepath)
        except Exception as e:
            raise Exception(f"Excel导出失败: {e}")
    
    @staticmethod
    def export_parts_list(parts_data, format='csv'):
        """导出配件列表"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format.lower() == 'csv':
            filename = f'配件列表_{timestamp}.csv'
            headers = ['配件ID', '配件名称', '配件编号', '类别', '品牌', '规格', '单位', 
                      '进价', '售价', '库存数量', '最小库存', '供应商']
            
            csv_data = []
            for part in parts_data:
                csv_data.append({
                    '配件ID': part.part_id,
                    '配件名称': part.part_name,
                    '配件编号': part.part_code,
                    '类别': part.category,
                    '品牌': part.brand,
                    '规格': part.specification,
                    '单位': part.unit,
                    '进价': part.purchase_price,
                    '售价': part.selling_price,
                    '库存数量': part.stock_quantity,
                    '最小库存': part.min_stock,
                    '供应商': part.supplier
                })
            
            return ExportUtils.export_to_csv(csv_data, filename, headers)
        
        elif format.lower() == 'json':
            filename = f'配件列表_{timestamp}.json'
            json_data = [part.to_dict() for part in parts_data]
            return ExportUtils.export_to_json(json_data, filename)
        
        else:
            raise ValueError(f"不支持的导出格式: {format}")
    
    @staticmethod
    def export_repair_orders(orders_data, format='csv'):
        """导出维修订单"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format.lower() == 'csv':
            filename = f'维修订单_{timestamp}.csv'
            headers = ['订单号', '客户姓名', '车辆类型', '车牌号', '维修日期', 
                      '故障描述', '维修内容', '工时费', '配件费', '总金额', '状态', '技师']
            
            csv_data = []
            for order in orders_data:
                csv_data.append({
                    '订单号': order.order_id,
                    '客户姓名': getattr(order, 'customer_name', ''),
                    '车辆类型': order.vehicle_type,
                    '车牌号': order.vehicle_number,
                    '维修日期': order.repair_date,
                    '故障描述': order.fault_description,
                    '维修内容': order.repair_content,
                    '工时费': order.labor_cost,
                    '配件费': order.parts_cost,
                    '总金额': order.total_amount,
                    '状态': order.status,
                    '技师': order.technician
                })
            
            return ExportUtils.export_to_csv(csv_data, filename, headers)
        
        elif format.lower() == 'json':
            filename = f'维修订单_{timestamp}.json'
            json_data = [order.to_dict() for order in orders_data]
            return ExportUtils.export_to_json(json_data, filename)
        
        else:
            raise ValueError(f"不支持的导出格式: {format}")
    
    @staticmethod
    def export_customers(customers_data, format='csv'):
        """导出客户列表"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format.lower() == 'csv':
            filename = f'客户列表_{timestamp}.csv'
            headers = ['客户ID', '客户姓名', '联系电话', '地址', '车辆信息']
            
            csv_data = []
            for customer in customers_data:
                csv_data.append({
                    '客户ID': customer.customer_id,
                    '客户姓名': customer.customer_name,
                    '联系电话': customer.phone,
                    '地址': customer.address,
                    '车辆信息': customer.vehicle_info
                })
            
            return ExportUtils.export_to_csv(csv_data, filename, headers)
        
        elif format.lower() == 'json':
            filename = f'客户列表_{timestamp}.json'
            json_data = [customer.to_dict() for customer in customers_data]
            return ExportUtils.export_to_json(json_data, filename)
        
        else:
            raise ValueError(f"不支持的导出格式: {format}")
    
    @staticmethod
    def export_report(report_data, report_name, format='json'):
        """导出报表数据"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{report_name}_{timestamp}.{format}'
        
        if format.lower() == 'json':
            return ExportUtils.export_to_json(report_data, filename)
        elif format.lower() == 'csv':
            # 如果是复杂的报表数据，需要特殊处理
            if isinstance(report_data, dict):
                # 将字典转换为列表格式
                csv_data = []
                for key, value in report_data.items():
                    csv_data.append({'项目': key, '值': value})
                return ExportUtils.export_to_csv(csv_data, filename)
            else:
                return ExportUtils.export_to_csv(report_data, filename)
        elif format.lower() == 'txt':
            return ExportUtils.export_to_txt(report_data, filename)
        else:
            raise ValueError(f"不支持的导出格式: {format}")
    
    @staticmethod
    def get_export_files():
        """获取导出文件列表"""
        try:
            if not REPORT_DIR.exists():
                return []
            
            export_files = []
            for file_path in REPORT_DIR.iterdir():
                if file_path.is_file():
                    stat = file_path.stat()
                    export_files.append({
                        'name': file_path.name,
                        'path': str(file_path),
                        'size': stat.st_size,
                        'create_time': datetime.fromtimestamp(stat.st_ctime),
                        'modify_time': datetime.fromtimestamp(stat.st_mtime)
                    })
            
            # 按修改时间倒序排列
            export_files.sort(key=lambda x: x['modify_time'], reverse=True)
            return export_files
        except Exception as e:
            raise Exception(f"获取导出文件列表失败: {e}")
    
    @staticmethod
    def delete_export_file(file_path):
        """删除导出文件"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return True
            return False
        except Exception as e:
            raise Exception(f"删除导出文件失败: {e}")